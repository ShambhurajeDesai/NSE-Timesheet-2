# # """
# # NSE Timesheet Automation System
# # ================================
# # FastAPI Endpoint: Biometric File Parser

# # Reads the NSE biometric Excel file (MusterInOutReport sheet)
# # and extracts structured attendance data per employee per day.

# # File Structure (per employee = 7 rows):
# #   Row 1 → ShiftCode   (one value per date column)
# #   Row 2 → Login
# #   Row 3 → Logout
# #   Row 4 → WorkHours
# #   Row 5 → ExtraWorkHours
# #   Row 6 → Status      (PP, PP*, AA, WW, WW*, HH, HH*, PA, PA*)
# #   Row 7 → Summary     (P=20, A=0, W=8, H=2, ...)

# # Usage:
# #   uvicorn biometric_parser:app --reload --port 8000
# #   POST http://localhost:8000/api/bio/parse  (with file upload)
# #   GET  http://localhost:8000/api/bio/parse-sample (uses hardcoded file path)
# # """



# # from fastapi import FastAPI, File, UploadFile, HTTPException, Query
# # from fastapi.responses import JSONResponse
# # from pydantic import BaseModel
# # from typing import Optional
# # import openpyxl
# # import pandas as pd
# # from io import BytesIO
# # from datetime import date
# # import re
# # import traceback

# # import calendar
# # from datetime import datetime

# # app = FastAPI(
# #     title="NSE Timesheet Automation API",
# #     description="Parses NSE biometric Excel files and extracts structured attendance records",
# #     version="1.0.0"
# # )


# # # ─────────────────────────────────────────────────────────────
# # # PYDANTIC MODELS (Response Schemas)
# # # ─────────────────────────────────────────────────────────────

# # class DailyAttendance(BaseModel):
# #     day: int                          # 1, 2, 3 ... 30
# #     shift_code: Optional[str]         # IO
# #     login: Optional[str]              # 09:32 or None
# #     logout: Optional[str]             # 18:32 or None
# #     work_hours: Optional[str]         # 9:00 or None
# #     extra_work_hours: Optional[str]   # 0:30 or None
# #     status: Optional[str]             # PP, PP*, AA, WW, HH, PA etc.
# #     timesheet_value: Optional[float]  # 1 / 0.5 / 0 / None (blank)


# # class EmployeeSummary(BaseModel):
# #     present: int      # P=20
# #     absent: int       # A=0
# #     weekend: int      # W=8
# #     holiday: int      # H=2
# #     overtime: int     # O=0
# #     flexi: int        # F=0
# #     late: int         # L=0
# #     comp_off: int     # C=0


# # class EmployeeRecord(BaseModel):
# #     emp_code: str                      # VT000010349
# #     emp_name: str                      # Sachin Bijwar
# #     po_number: str                     # PO-NSEIL OU-26-2027
# #     summary: EmployeeSummary
# #     attendance: list[DailyAttendance]  # one per requested day


# # class BiometricParseResponse(BaseModel):
# #     month_range: str                   # From :01/04/2026 to 30/04/2026
# #     total_employees: int
# #     days_extracted: list[int]          # [1, 2, 3, 4, 5]
# #     employees: list[EmployeeRecord]


# # # ─────────────────────────────────────────────────────────────
# # # CORE PARSER
# # # ─────────────────────────────────────────────────────────────

# # def parse_biometric_excel(file_bytes: bytes, days_filter: list[int]) -> dict:
# #     """
# #     Main parser for the NSE biometric Excel file.

# #     Args:
# #         file_bytes   : Raw bytes of the uploaded .xlsx file
# #         days_filter  : List of day numbers to extract e.g. [1, 2, 3, 4, 5]

# #     Returns:
# #         Structured dict ready to be returned as JSON
# #     """

# #     wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
# #     ws = wb.active  # Sheet: MusterInOutReport

# #     # ── Step 1: Read the date header row (Row 6) ────────────────────
# #     # Row 6 layout: [EmpCode, EmpName, PONumber, '-', None, 1, 2, 3, ...]
# #     # Date numbers (1–30) are scattered with None gaps for merged cells
# #     header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]

# #     # Build map: day_number → column_index (0-based)
# #     # Only pick first occurrence of each day number (1–30)
# #     date_col_map = {}
# #     for col_idx, cell_val in enumerate(header_row):
# #         if isinstance(cell_val, int) and 1 <= cell_val <= 30:
# #             if cell_val not in date_col_map:
# #                 date_col_map[cell_val] = col_idx

# #     # Validate requested days exist in this file
# #     available_days = sorted(date_col_map.keys())
# #     invalid_days = [d for d in days_filter if d not in date_col_map]
# #     if invalid_days:
# #         raise ValueError(
# #             f"Days {invalid_days} not found in biometric file. "
# #             f"Available days: {available_days}"
# #         )

# #     # ── Step 2: Read month range from Row 5 ─────────────────────────
# #     row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
# #     month_range = str(row5[0]) if row5[0] else "Unknown"

# #     # ── Step 3: Read all employee rows (Row 7 onwards) ───────────────
# #     # Each employee = exactly 7 rows:
# #     #   ShiftCode | Login | Logout | WorkHours | ExtraWorkHours | Status | Summary
# #     ROWS_PER_EMPLOYEE = 7
# #     ROW_TYPES = ["ShiftCode", "Login", "Logout", "WorkHours", "ExtraWorkHours", "Status", "Summary"]

# #     # Load all data rows into memory at once (bulk read — no N+1)
# #     all_data_rows = list(ws.iter_rows(min_row=7, values_only=True))
# #     wb.close()

# #     employees = []
# #     i = 0

# #     while i < len(all_data_rows):
# #         # ── Identify the start of an employee block ──────────────────
# #         # The ShiftCode row always has the VT ID in column 0
# #         shift_row = all_data_rows[i]

# #         emp_code = shift_row[0]
# #         if not emp_code or not str(emp_code).startswith("VT"):
# #             i += 1
# #             continue  # Skip any non-employee rows

# #         # ── Safety check: ensure we have all 7 rows ──────────────────
# #         if i + ROWS_PER_EMPLOYEE > len(all_data_rows):
# #             break

# #         # ── Extract all 7 rows for this employee ─────────────────────
# #         block = all_data_rows[i : i + ROWS_PER_EMPLOYEE]

# #         # Map row type → raw row data (list of cell values)
# #         row_map = {row_type: block[j] for j, row_type in enumerate(ROW_TYPES)}

# #         # ── Extract static fields ─────────────────────────────────────
# #         emp_name  = str(shift_row[1]).strip() if shift_row[1] else ""
# #         po_number = str(shift_row[2]).strip() if shift_row[2] else ""

# #         # ── Parse Summary row ─────────────────────────────────────────
# #         # Summary row values: ['P=20', 'A=0', 'W=8', 'H=2', 'O=0', 'F=0', 'L=0', 'C=0']
# #         # These are spread across the date columns — we read non-None values
# #         summary_row = row_map["Summary"]
# #         summary_values = [
# #             str(v).strip() for v in summary_row if v is not None and str(v).strip() != ""
# #         ]
# #         summary = parse_summary(summary_values)

# #         # ── Extract daily attendance for requested days only ──────────
# #         attendance = []
# #         for day in sorted(days_filter):
# #             col_idx = date_col_map[day]

# #             raw_shift      = clean_val(row_map["ShiftCode"][col_idx])
# #             raw_login      = clean_val(row_map["Login"][col_idx])
# #             raw_logout     = clean_val(row_map["Logout"][col_idx])
# #             raw_workhours  = clean_val(row_map["WorkHours"][col_idx])
# #             raw_extrahours = clean_val(row_map["ExtraWorkHours"][col_idx])
# #             raw_status     = clean_val(row_map["Status"][col_idx])

# #             # Determine timesheet value from status + work hours
# #             ts_value = map_status_to_timesheet_value(raw_status, raw_workhours)

# #             attendance.append(DailyAttendance(
# #                 day=day,
# #                 shift_code=raw_shift,
# #                 login=raw_login,
# #                 logout=raw_logout,
# #                 work_hours=raw_workhours,
# #                 extra_work_hours=raw_extrahours,
# #                 status=raw_status,
# #                 timesheet_value=ts_value
# #             ))

# #         employees.append(EmployeeRecord(
# #             emp_code=str(emp_code).strip(),
# #             emp_name=emp_name,
# #             po_number=po_number,
# #             summary=summary,
# #             attendance=attendance
# #         ))

# #         i += ROWS_PER_EMPLOYEE  # Jump to next employee block

# #     return BiometricParseResponse(
# #         month_range=month_range,
# #         total_employees=len(employees),
# #         days_extracted=sorted(days_filter),
# #         employees=employees
# #     )


# # # ─────────────────────────────────────────────────────────────
# # # HELPER FUNCTIONS
# # # ─────────────────────────────────────────────────────────────

# # def clean_val(val) -> Optional[str]:
# #     """
# #     Converts cell value to clean string.
# #     Returns None for empty/dash values.
# #     """
# #     if val is None:
# #         return None
# #     s = str(val).strip()
# #     if s in ("-", "", "None"):
# #         return None
# #     return s


# # def parse_summary(summary_values: list[str]) -> EmployeeSummary:
# #     """
# #     Parses summary tokens like ['P=20', 'A=0', 'W=8', 'H=2', 'O=0', 'F=0', 'L=0', 'C=0']
# #     into a structured EmployeeSummary object.
# #     """
# #     summary_map = {}
# #     for token in summary_values:
# #         if "=" in token:
# #             key, _, val = token.partition("=")
# #             try:
# #                 summary_map[key.strip().upper()] = int(val.strip())
# #             except ValueError:
# #                 summary_map[key.strip().upper()] = 0

# #     return EmployeeSummary(
# #         present  = summary_map.get("P", 0),
# #         absent   = summary_map.get("A", 0),
# #         weekend  = summary_map.get("W", 0),
# #         holiday  = summary_map.get("H", 0),
# #         overtime = summary_map.get("O", 0),
# #         flexi    = summary_map.get("F", 0),
# #         late     = summary_map.get("L", 0),
# #         comp_off = summary_map.get("C", 0),
# #     )


# # def map_status_to_timesheet_value(status: Optional[str], work_hours: Optional[str]) -> Optional[float]:
# #     """
# #     Maps biometric status code → timesheet cell value.

# #     Returns:
# #       1.0   → Full day present
# #       0.5   → Half day
# #       0.0   → Absent on working day
# #       None  → Weekend or Holiday (blank in timesheet)
# #     """
# #     if status is None:
# #         return None

# #     status = status.upper().strip()

# #     # Full present
# #     if status in ("PP", "PP*"):
# #         # Extra validation: if work hours provided, check >= 9hrs
# #         if work_hours:
# #             hours_decimal = parse_work_hours(work_hours)
# #             if hours_decimal is not None and hours_decimal >= 9.0:
# #                 return 1.0
# #             elif hours_decimal is not None and hours_decimal >= 4.5:
# #                 return 0.5
# #             else:
# #                 return 1.0  # trust status if hours parsing fails
# #         return 1.0

# #     # Half day
# #     if status in ("PA", "PA*"):
# #         return 0.5

# #     # Absent
# #     if status == "AA":
# #         return 0.0

# #     # Weekend (blank in timesheet)
# #     if status in ("WW", "WW*", "WW#"):
# #         return None

# #     # Holiday (blank in timesheet)
# #     if status in ("HH", "HH#"):
# #         return None

# #     # Worked on holiday → 1
# #     if status == "HH*":
# #         return 1.0

# #     # Unknown status → None
# #     return None


# # def parse_work_hours(work_hours_str: str) -> Optional[float]:
# #     """
# #     Converts '9:30' → 9.5 (hours as decimal float).
# #     Returns None if parsing fails.
# #     """
# #     if not work_hours_str or work_hours_str == "-":
# #         return None
# #     try:
# #         parts = work_hours_str.split(":")
# #         hours = int(parts[0])
# #         minutes = int(parts[1]) if len(parts) > 1 else 0
# #         return round(hours + minutes / 60, 2)
# #     except Exception:
# #         return None


# # def get_month_details(file_bytes: bytes):
# #     """
# #     Reads the 'From : dd/mm/yyyy to dd/mm/yyyy' text from the Excel sheet.

# #     Returns:
# #         {
# #             "month": "April",
# #             "month_number": 4,
# #             "year": 2026,
# #             "days_in_month": 30,
# #             "start_date": "01/04/2026",
# #             "end_date": "30/04/2026"
# #         }
# #     """

# #     df = pd.read_excel(
# #         BytesIO(file_bytes),
# #         header=None,
# #         engine="openpyxl"
# #     )

# #     date_text = None

# #     # Search entire sheet instead of assuming row 5
# #     for row in df.itertuples(index=False):
# #         for cell in row:
# #             if pd.notna(cell):
# #                 text = str(cell)

# #                 if "From" in text and "to" in text:
# #                     date_text = text
# #                     break

# #         if date_text:
# #             break

# #     if not date_text:
# #         raise ValueError("Unable to locate attendance period in Excel.")

# #     match = re.search(
# #         r'(\d{2}/\d{2}/\d{4})\s*to\s*(\d{2}/\d{2}/\d{4})',
# #         date_text
# #     )

# #     if not match:
# #         raise ValueError("Invalid attendance date format.")

# #     start_date = datetime.strptime(match.group(1), "%d/%m/%Y")
# #     end_date = datetime.strptime(match.group(2), "%d/%m/%Y")

# #     return {
# #         "month": start_date.strftime("%B"),
# #         "month_number": start_date.month,
# #         "year": start_date.year,
# #         "days_in_month": calendar.monthrange(
# #             start_date.year,
# #             start_date.month
# #         )[1],
# #         "start_date": start_date.strftime("%d/%m/%Y"),
# #         "end_date": end_date.strftime("%d/%m/%Y")
# #     }






# # # ─────────────────────────────────────────────────────────────
# # # API ENDPOINTS
# # # ─────────────────────────────────────────────────────────────

# # @app.get("/", tags=["Health"])
# # async def root():
# #     return {
# #         "service": "NSE Timesheet Automation API",
# #         "version": "1.0.0",
# #         "status": "running",
# #         "endpoints": {
# #             "upload_and_parse": "POST /api/bio/parse",
# #             "parse_sample":     "GET  /api/bio/parse-sample",
# #             "docs":             "GET  /docs"
# #         }
# #     }


# # @app.get("/health", tags=["Health"])
# # async def health_check():
# #     return {"status": "ok"}


# # # ── ENDPOINT 1: Upload biometric file ───────────────────────────────

# # @app.post(
# #     "/api/bio/parse",
# #     response_model=BiometricParseResponse,
# #     tags=["Biometric"],
# #     summary="Upload & Parse Biometric File",
# #     description="""
# #     Upload an NSE biometric Excel file (.xlsx) and extract attendance records.

# #     **How to use in Postman:**
# #     - Method: POST
# #     - URL: http://localhost:8000/api/bio/parse
# #     - Body → form-data:
# #         - Key: `file`  Type: File  Value: (select your .xlsx file)
# #         - Key: `days`  Type: Text  Value: `1,2,3,4,5`  (optional, default: 1-5)

# #     **Returns:** Structured JSON with one record per employee per requested day.
# #     """
# # )
# # async def parse_biometric_file(
# #     file: UploadFile = File(..., description="NSE biometric .xlsx file"),
# #     days: str = Query(
# #         default="1,2,3,4,5",
# #         description="Comma-separated day numbers to extract. Example: 1,2,3,4,5 or 1,2,3"
# #     )
# # ):
# #     # ── Validate file type ───────────────────────────────────────────
# #     if not file.filename.endswith((".xlsx", ".xls")):
# #         raise HTTPException(
# #             status_code=400,
# #             detail=f"Invalid file type '{file.filename}'. Only .xlsx or .xls files are accepted."
# #         )

# #     # ── Parse requested days ─────────────────────────────────────────
# #     try:
# #         days_list = [int(d.strip()) for d in days.split(",") if d.strip()]
# #         if not days_list:
# #             raise ValueError("Empty days list")
# #         if any(d < 1 or d > 31 for d in days_list):
# #             raise ValueError("Day numbers must be between 1 and 31")
# #         days_list = list(set(days_list))  # remove duplicates
# #     except ValueError as e:
# #         raise HTTPException(
# #             status_code=400,
# #             detail=f"Invalid 'days' parameter: {str(e)}. Expected format: '1,2,3,4,5'"
# #         )

# #     # ── Read file bytes ──────────────────────────────────────────────
# #     try:
# #         file_bytes = await file.read()
# #         if len(file_bytes) == 0:
# #             raise HTTPException(status_code=400, detail="Uploaded file is empty.")
# #     except Exception as e:
# #         raise HTTPException(status_code=400, detail=f"Failed to read uploaded file: {str(e)}")

# #     # ── Parse ────────────────────────────────────────────────────────
# #     try:
# #         result = parse_biometric_excel(file_bytes, days_list)
# #         return result
# #     except ValueError as e:
# #         raise HTTPException(status_code=422, detail=str(e))
# #     except Exception as e:
# #         raise HTTPException(
# #             status_code=500,
# #             detail=f"Parsing failed: {str(e)}\n{traceback.format_exc()}"
# #         )


# # # ── ENDPOINT 2: Parse from server path (for quick Postman testing) ───

# # @app.get(
# #     "/api/bio/parse-sample",
# #     response_model=BiometricParseResponse,
# #     tags=["Biometric"],
# #     summary="Parse Sample File (Dev/Test Only)",
# #     description="""
# #     Parses a hardcoded biometric file from the server filesystem.
# #     Useful for quick Postman testing without needing to upload a file.

# #     **How to use in Postman:**
# #     - Method: GET
# #     - URL: http://localhost:8000/api/bio/parse-sample?days=1,2,3,4,5

# #     **Query params:**
# #     - `days` → comma-separated day numbers (default: 1,2,3,4,5)
# #     - `emp_code` → filter by specific VT ID (optional)
# #     """
# # )
# # async def parse_sample_file(
# #     days: str = Query(
# #         default="1,2,3,4,5",
# #         description="Comma-separated day numbers. Example: 1,2,3,4,5"
# #     ),
# #     emp_code: Optional[str] = Query(
# #         default=None,
# #         description="Filter by employee VT ID. Example: VT000010349"
# #     )
# # ):
# #     # Hardcoded sample file path (update this to your actual path)
# #     SAMPLE_FILE_PATH = "NSE_biometric_data.xlsx"

# #     # ── Parse requested days ─────────────────────────────────────────
# #     try:
# #         days_list = [int(d.strip()) for d in days.split(",") if d.strip()]
# #         if not days_list:
# #             raise ValueError("Empty days list")
# #         if any(d < 1 or d > 31 for d in days_list):
# #             raise ValueError("Day numbers must be between 1 and 31")
# #         days_list = list(set(days_list))
# #     except ValueError as e:
# #         raise HTTPException(
# #             status_code=400,
# #             detail=f"Invalid 'days' parameter: {str(e)}"
# #         )

# #     # ── Read file ────────────────────────────────────────────────────
# #     try:
# #         with open(SAMPLE_FILE_PATH, "rb") as f:
# #             file_bytes = f.read()
# #     except FileNotFoundError:
# #         raise HTTPException(
# #             status_code=404,
# #             detail=f"Sample file not found at '{SAMPLE_FILE_PATH}'. "
# #                    f"Use POST /api/bio/parse to upload your file instead."
# #         )

# #     # ── Parse ────────────────────────────────────────────────────────
# #     try:
# #         result = parse_biometric_excel(file_bytes, days_list)

# #         # Optional: filter by emp_code
# #         if emp_code:
# #             filtered = [e for e in result.employees if e.emp_code == emp_code.strip()]
# #             if not filtered:
# #                 raise HTTPException(
# #                     status_code=404,
# #                     detail=f"Employee '{emp_code}' not found in biometric file."
# #                 )
# #             result.employees = filtered
# #             result.total_employees = len(filtered)

# #         return result
# #     except HTTPException:
# #         raise
# #     except ValueError as e:
# #         raise HTTPException(status_code=422, detail=str(e))
# #     except Exception as e:
# #         raise HTTPException(
# #             status_code=500,
# #             detail=f"Parsing failed: {str(e)}\n{traceback.format_exc()}"
# #         )


# # # ── ENDPOINT 3: Get single employee attendance ───────────────────────

# # @app.post(
# #     "/api/bio/parse/employee/{emp_code}",
# #     response_model=EmployeeRecord,
# #     tags=["Biometric"],
# #     summary="Get Single Employee Attendance",
# #     description="""
# #     Upload a biometric file and extract attendance for a specific employee only.

# #     **How to use in Postman:**
# #     - Method: POST
# #     - URL: http://localhost:8000/api/bio/parse/employee/VT000010349?days=1,2,3,4,5
# #     - Body → form-data:
# #         - Key: `file`  Type: File  Value: (select your .xlsx file)
# #     """
# # )
# # async def parse_single_employee(
# #     emp_code: str,
# #     file: UploadFile = File(...),
# #     days: str = Query(default="1,2,3,4,5")
# # ):
# #     if not file.filename.endswith((".xlsx", ".xls")):
# #         raise HTTPException(status_code=400, detail="Only .xlsx or .xls files accepted.")

# #     try:
# #         days_list = [int(d.strip()) for d in days.split(",") if d.strip()]
# #     except ValueError:
# #         raise HTTPException(status_code=400, detail="Invalid days format. Use: 1,2,3,4,5")

# #     file_bytes = await file.read()

# #     try:
# #         result = parse_biometric_excel(file_bytes, days_list)
# #     except Exception as e:
# #         raise HTTPException(status_code=500, detail=str(e))

# #     # Find the specific employee
# #     emp_record = next(
# #         (e for e in result.employees if e.emp_code == emp_code.strip()),
# #         None
# #     )

# #     if not emp_record:
# #         raise HTTPException(
# #             status_code=404,
# #             detail=f"Employee '{emp_code}' not found. "
# #                    f"Total employees in file: {result.total_employees}"
# #         )

# #     return emp_record


# # # ─────────────────────────────────────────────────────────────
# # # RUN SERVER
# # # ─────────────────────────────────────────────────────────────

# # if __name__ == "__main__":
# #     import uvicorn
# #     uvicorn.run(
# #         "biometric_parser:app",
# #         host="0.0.0.0",
# #         port=8000,
# #         reload=True,         # Auto-reload on file change (dev mode)
# #         log_level="info"
# #     )


# """
# NSE Timesheet Automation System
# ================================
# FastAPI Endpoint: Biometric File Parser
 
# Reads the NSE biometric Excel file (MusterInOutReport sheet)
# and extracts structured attendance data per employee per day.
 
# File Structure (per employee = 7 rows):
#   Row 1 → ShiftCode   (one value per date column)
#   Row 2 → Login
#   Row 3 → Logout
#   Row 4 → WorkHours
#   Row 5 → ExtraWorkHours
#   Row 6 → Status      (PP, PP*, AA, WW, WW*, HH, HH*, PA, PA*)
#   Row 7 → Summary     (P=20, A=0, W=8, H=2, ...)
 
# Days handling:
#   By default, `days` is NOT hardcoded. The API reads the "From : dd/mm/yyyy
#   to dd/mm/yyyy" text embedded in the sheet, figures out the month/year,
#   and uses calendar.monthrange() to determine exactly how many days that
#   month has (28/29/30/31). It then extracts ALL of those days automatically.
#   You can still override this by explicitly passing `days=1,2,3` etc.
 
# Usage:
#   uvicorn biometric_parser:app --reload --port 8000
#   POST http://localhost:8000/api/bio/parse  (with file upload)
#   GET  http://localhost:8000/api/bio/parse-sample (uses hardcoded file path)
# """
 
 
# from fastapi import FastAPI, File, UploadFile, HTTPException, Query
# from fastapi.responses import JSONResponse
# from pydantic import BaseModel
# from typing import Optional
# import openpyxl
# import pandas as pd
# from io import BytesIO
# from datetime import date
# import re
# import traceback
 
# import calendar
# from datetime import datetime
 
# app = FastAPI(
#     title="NSE Timesheet Automation API",
#     description="Parses NSE biometric Excel files and extracts structured attendance records",
#     version="1.0.0"
# )
 
 
# # ─────────────────────────────────────────────────────────────
# # PYDANTIC MODELS (Response Schemas)
# # ─────────────────────────────────────────────────────────────
 
# class DailyAttendance(BaseModel):
#     day: int                          # 1, 2, 3 ... 31
#     shift_code: Optional[str]         # IO
#     login: Optional[str]              # 09:32 or None
#     logout: Optional[str]             # 18:32 or None
#     work_hours: Optional[str]         # 9:00 or None
#     extra_work_hours: Optional[str]   # 0:30 or None
#     status: Optional[str]             # PP, PP*, AA, WW, HH, PA etc.
#     timesheet_value: Optional[float]  # 1 / 0.5 / 0 / None (blank)
 
 
# class EmployeeSummary(BaseModel):
#     present: int      # P=20
#     absent: int       # A=0
#     weekend: int      # W=8
#     holiday: int      # H=2
#     overtime: int     # O=0
#     flexi: int        # F=0
#     late: int         # L=0
#     comp_off: int     # C=0
 
 
# class EmployeeRecord(BaseModel):
#     emp_code: str                      # VT000010349
#     emp_name: str                      # Sachin Bijwar
#     po_number: str                     # PO-NSEIL OU-26-2027
#     summary: EmployeeSummary
#     attendance: list[DailyAttendance]  # one per requested day
 
 
# class MonthInfo(BaseModel):
#     month: str            # "April"
#     month_number: int     # 4
#     year: int             # 2026
#     days_in_month: int    # 30
#     start_date: str       # "01/04/2026"
#     end_date: str         # "30/04/2026"
 
 
# class BiometricParseResponse(BaseModel):
#     month_range: str                   # From :01/04/2026 to 30/04/2026
#     month_info: Optional[MonthInfo]    # structured month detection result
#     total_employees: int
#     days_extracted: list[int]          # [1, 2, ... days_in_month] (auto) or user override
#     employees: list[EmployeeRecord]
 
 
# # ─────────────────────────────────────────────────────────────
# # MONTH / DAY DETECTION
# # ─────────────────────────────────────────────────────────────
 
# def get_month_details(file_bytes: bytes) -> dict:
#     """
#     Reads the 'From : dd/mm/yyyy to dd/mm/yyyy' text from the Excel sheet.
 
#     Returns:
#         {
#             "month": "April",
#             "month_number": 4,
#             "year": 2026,
#             "days_in_month": 30,
#             "start_date": "01/04/2026",
#             "end_date": "30/04/2026"
#         }
#     """
 
#     df = pd.read_excel(
#         BytesIO(file_bytes),
#         header=None,
#         engine="openpyxl"
#     )
 
#     date_text = None
 
#     # Search entire sheet instead of assuming row 5
#     for row in df.itertuples(index=False):
#         for cell in row:
#             if pd.notna(cell):
#                 text = str(cell)
 
#                 if "From" in text and "to" in text:
#                     date_text = text
#                     break
 
#         if date_text:
#             break
 
#     if not date_text:
#         raise ValueError("Unable to locate attendance period in Excel.")
 
#     match = re.search(
#         r'(\d{2}/\d{2}/\d{4})\s*to\s*(\d{2}/\d{2}/\d{4})',
#         date_text
#     )
 
#     if not match:
#         raise ValueError("Invalid attendance date format.")
 
#     start_date = datetime.strptime(match.group(1), "%d/%m/%Y")
#     end_date = datetime.strptime(match.group(2), "%d/%m/%Y")
 
#     return {
#         "month": start_date.strftime("%B"),
#         "month_number": start_date.month,
#         "year": start_date.year,
#         "days_in_month": calendar.monthrange(
#             start_date.year,
#             start_date.month
#         )[1],
#         "start_date": start_date.strftime("%d/%m/%Y"),
#         "end_date": end_date.strftime("%d/%m/%Y")
#     }
 
 
# def get_all_days_for_month(file_bytes: bytes) -> tuple[list[int], dict]:
#     """
#     Detects the month/year embedded in the sheet and returns every valid
#     day number for that month (1 .. days_in_month) using calendar.monthrange.
 
#     Returns:
#         (days_list, month_details_dict)
#         e.g. ([1, 2, 3, ..., 30], {"month": "April", "year": 2026, ...})
#     """
#     month_details = get_month_details(file_bytes)
#     days_in_month = month_details["days_in_month"]
#     days_list = list(range(1, days_in_month + 1))
#     return days_list, month_details
 
 
# # ─────────────────────────────────────────────────────────────
# # CORE PARSER
# # ─────────────────────────────────────────────────────────────
 
# def parse_biometric_excel(file_bytes: bytes, days_filter: list[int]) -> dict:
#     """
#     Main parser for the NSE biometric Excel file.
 
#     Args:
#         file_bytes   : Raw bytes of the uploaded .xlsx file
#         days_filter  : List of day numbers to extract e.g. [1, 2, ..., 30]
#                        (typically produced automatically via
#                        get_all_days_for_month, unless the caller overrides it)
 
#     Returns:
#         Structured dict ready to be returned as JSON
#     """
 
#     # Try to detect month info up front. If this fails (e.g. the period
#     # text isn't present in some non-standard export) we degrade gracefully
#     # and just proceed with whatever days_filter was supplied.
#     month_info = None
#     try:
#         month_info = get_month_details(file_bytes)
#     except ValueError:
#         pass
 
#     wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
#     ws = wb.active  # Sheet: MusterInOutReport
 
#     # ── Step 1: Read the date header row (Row 6) ────────────────────
#     # Row 6 layout: [EmpCode, EmpName, PONumber, '-', None, 1, 2, 3, ...]
#     # Date numbers (1–31) are scattered with None gaps for merged cells
#     header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
 
#     # Build map: day_number → column_index (0-based)
#     # Only pick first occurrence of each day number (1–31)
#     # NOTE: upper bound raised from 30 → 31 so 31-day months aren't truncated.
#     date_col_map = {}
#     for col_idx, cell_val in enumerate(header_row):
#         if isinstance(cell_val, int) and 1 <= cell_val <= 31:
#             if cell_val not in date_col_map:
#                 date_col_map[cell_val] = col_idx
 
#     # Validate requested days exist in this file
#     available_days = sorted(date_col_map.keys())
#     invalid_days = [d for d in days_filter if d not in date_col_map]
#     if invalid_days:
#         raise ValueError(
#             f"Days {invalid_days} not found in biometric file. "
#             f"Available days: {available_days}"
#         )
 
#     # ── Step 2: Read month range from Row 5 ─────────────────────────
#     row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
#     month_range = str(row5[0]) if row5[0] else "Unknown"
 
#     # ── Step 3: Read all employee rows (Row 7 onwards) ───────────────
#     # Each employee = exactly 7 rows:
#     #   ShiftCode | Login | Logout | WorkHours | ExtraWorkHours | Status | Summary
#     ROWS_PER_EMPLOYEE = 7
#     ROW_TYPES = ["ShiftCode", "Login", "Logout", "WorkHours", "ExtraWorkHours", "Status", "Summary"]
 
#     # Load all data rows into memory at once (bulk read — no N+1)
#     all_data_rows = list(ws.iter_rows(min_row=7, values_only=True))
#     wb.close()
 
#     employees = []
#     i = 0
 
#     while i < len(all_data_rows):
#         # ── Identify the start of an employee block ──────────────────
#         # The ShiftCode row always has the VT ID in column 0
#         shift_row = all_data_rows[i]
 
#         emp_code = shift_row[0]
#         if not emp_code or not str(emp_code).startswith("VT"):
#             i += 1
#             continue  # Skip any non-employee rows
 
#         # ── Safety check: ensure we have all 7 rows ──────────────────
#         if i + ROWS_PER_EMPLOYEE > len(all_data_rows):
#             break
 
#         # ── Extract all 7 rows for this employee ─────────────────────
#         block = all_data_rows[i : i + ROWS_PER_EMPLOYEE]
 
#         # Map row type → raw row data (list of cell values)
#         row_map = {row_type: block[j] for j, row_type in enumerate(ROW_TYPES)}
 
#         # ── Extract static fields ─────────────────────────────────────
#         emp_name  = str(shift_row[1]).strip() if shift_row[1] else ""
#         po_number = str(shift_row[2]).strip() if shift_row[2] else ""
 
#         # ── Parse Summary row ─────────────────────────────────────────
#         # Summary row values: ['P=20', 'A=0', 'W=8', 'H=2', 'O=0', 'F=0', 'L=0', 'C=0']
#         # These are spread across the date columns — we read non-None values
#         summary_row = row_map["Summary"]
#         summary_values = [
#             str(v).strip() for v in summary_row if v is not None and str(v).strip() != ""
#         ]
#         summary = parse_summary(summary_values)
 
#         # ── Extract daily attendance for requested days only ──────────
#         attendance = []
#         for day in sorted(days_filter):
#             col_idx = date_col_map[day]
 
#             raw_shift      = clean_val(row_map["ShiftCode"][col_idx])
#             raw_login      = clean_val(row_map["Login"][col_idx])
#             raw_logout     = clean_val(row_map["Logout"][col_idx])
#             raw_workhours  = clean_val(row_map["WorkHours"][col_idx])
#             raw_extrahours = clean_val(row_map["ExtraWorkHours"][col_idx])
#             raw_status     = clean_val(row_map["Status"][col_idx])
 
#             # Determine timesheet value from status + work hours
#             ts_value = map_status_to_timesheet_value(raw_status, raw_workhours)
 
#             attendance.append(DailyAttendance(
#                 day=day,
#                 shift_code=raw_shift,
#                 login=raw_login,
#                 logout=raw_logout,
#                 work_hours=raw_workhours,
#                 extra_work_hours=raw_extrahours,
#                 status=raw_status,
#                 timesheet_value=ts_value
#             ))
 
#         employees.append(EmployeeRecord(
#             emp_code=str(emp_code).strip(),
#             emp_name=emp_name,
#             po_number=po_number,
#             summary=summary,
#             attendance=attendance
#         ))
 
#         i += ROWS_PER_EMPLOYEE  # Jump to next employee block
 
#     return BiometricParseResponse(
#         month_range=month_range,
#         month_info=MonthInfo(**month_info) if month_info else None,
#         total_employees=len(employees),
#         days_extracted=sorted(days_filter),
#         employees=employees
#     )
 
 
# # ─────────────────────────────────────────────────────────────
# # HELPER FUNCTIONS
# # ─────────────────────────────────────────────────────────────
 
# def clean_val(val) -> Optional[str]:
#     """
#     Converts cell value to clean string.
#     Returns None for empty/dash values.
#     """
#     if val is None:
#         return None
#     s = str(val).strip()
#     if s in ("-", "", "None"):
#         return None
#     return s
 
 
# def parse_summary(summary_values: list[str]) -> EmployeeSummary:
#     """
#     Parses summary tokens like ['P=20', 'A=0', 'W=8', 'H=2', 'O=0', 'F=0', 'L=0', 'C=0']
#     into a structured EmployeeSummary object.
#     """
#     summary_map = {}
#     for token in summary_values:
#         if "=" in token:
#             key, _, val = token.partition("=")
#             try:
#                 summary_map[key.strip().upper()] = int(val.strip())
#             except ValueError:
#                 summary_map[key.strip().upper()] = 0
 
#     return EmployeeSummary(
#         present  = summary_map.get("P", 0),
#         absent   = summary_map.get("A", 0),
#         weekend  = summary_map.get("W", 0),
#         holiday  = summary_map.get("H", 0),
#         overtime = summary_map.get("O", 0),
#         flexi    = summary_map.get("F", 0),
#         late     = summary_map.get("L", 0),
#         comp_off = summary_map.get("C", 0),
#     )
 
 
# def map_status_to_timesheet_value(status: Optional[str], work_hours: Optional[str]) -> Optional[float]:
#     """
#     Maps biometric status code → timesheet cell value.
 
#     Returns:
#       1.0   → Full day present
#       0.5   → Half day
#       0.0   → Absent on working day
#       None  → Weekend or Holiday (blank in timesheet)
#     """
#     if status is None:
#         return None
 
#     status = status.upper().strip()
 
#     # Full present
#     if status in ("PP", "PP*"):
#         # Extra validation: if work hours provided, check >= 9hrs
#         if work_hours:
#             hours_decimal = parse_work_hours(work_hours)
#             if hours_decimal is not None and hours_decimal >= 9.0:
#                 return 1.0
#             elif hours_decimal is not None and hours_decimal >= 4.5:
#                 return 0.5
#             else:
#                 return 1.0  # trust status if hours parsing fails
#         return 1.0
 
#     # Half day
#     if status in ("PA", "PA*"):
#         return 0.5
 
#     # Absent
#     if status == "AA":
#         return 0.0
 
#     # Weekend (blank in timesheet)
#     if status in ("WW", "WW*", "WW#"):
#         return None
 
#     # Holiday (blank in timesheet)
#     if status in ("HH", "HH#"):
#         return None
 
#     # Worked on holiday → 1
#     if status == "HH*":
#         return 1.0
 
#     # Unknown status → None
#     return None
 
 
# def parse_work_hours(work_hours_str: str) -> Optional[float]:
#     """
#     Converts '9:30' → 9.5 (hours as decimal float).
#     Returns None if parsing fails.
#     """
#     if not work_hours_str or work_hours_str == "-":
#         return None
#     try:
#         parts = work_hours_str.split(":")
#         hours = int(parts[0])
#         minutes = int(parts[1]) if len(parts) > 1 else 0
#         return round(hours + minutes / 60, 2)
#     except Exception:
#         return None
 
 
# def resolve_days_list(file_bytes: bytes, days: Optional[str]) -> list[int]:
#     """
#     Decides which days to extract:
#       - If `days` is None/blank  → auto-detect month from the sheet and
#         return every day in that month (1 .. days_in_month) via
#         calendar.monthrange.
#       - If `days` is provided    → parse and validate the explicit override,
#         same as before.
#     """
#     if days is None or not days.strip():
#         try:
#             days_list, _ = get_all_days_for_month(file_bytes)
#             return days_list
#         except ValueError as e:
#             # Could not detect the month from the sheet — fall back to a
#             # conservative default rather than failing outright.
#             raise ValueError(
#                 f"Could not auto-detect month from file to determine days: {e}. "
#                 f"Pass an explicit 'days' query param (e.g. '1,2,3,...,30') instead."
#             )
 
#     days_list = [int(d.strip()) for d in days.split(",") if d.strip()]
#     if not days_list:
#         raise ValueError("Empty days list")
#     if any(d < 1 or d > 31 for d in days_list):
#         raise ValueError("Day numbers must be between 1 and 31")
#     return list(set(days_list))  # remove duplicates
 
 
# # ─────────────────────────────────────────────────────────────
# # API ENDPOINTS
# # ─────────────────────────────────────────────────────────────
 
# @app.get("/", tags=["Health"])
# async def root():
#     return {
#         "service": "NSE Timesheet Automation API",
#         "version": "1.0.0",
#         "status": "running",
#         "endpoints": {
#             "upload_and_parse": "POST /api/bio/parse",
#             "parse_sample":     "GET  /api/bio/parse-sample",
#             "docs":             "GET  /docs"
#         }
#     }
 
 
# @app.get("/health", tags=["Health"])
# async def health_check():
#     return {"status": "ok"}
 
 
# # ── ENDPOINT 1: Upload biometric file ───────────────────────────────
 
# @app.post(
#     "/api/bio/parse",
#     response_model=BiometricParseResponse,
#     tags=["Biometric"],
#     summary="Upload & Parse Biometric File",
#     description="""
#     Upload an NSE biometric Excel file (.xlsx) and extract attendance records.
 
#     **Days parameter (auto by default):**
#     If you don't pass `days`, the API reads the "From ... to ..." period
#     embedded in the sheet, detects the month/year, and uses
#     `calendar.monthrange()` to extract EVERY day in that month automatically
#     (28/29/30/31 days, whatever applies). Pass `days` explicitly only if you
#     want a subset.
 
#     **How to use in Postman:**
#     - Method: POST
#     - URL: http://localhost:8000/api/bio/parse
#     - Body → form-data:
#         - Key: `file`  Type: File  Value: (select your .xlsx file)
#         - Key: `days`  Type: Text  Value: `1,2,3,4,5`  (optional — omit for full month)
 
#     **Returns:** Structured JSON with one record per employee per requested day.
#     """
# )
# async def parse_biometric_file(
#     file: UploadFile = File(..., description="NSE biometric .xlsx file"),
#     days: Optional[str] = Query(
#         default=None,
#         description=(
#             "Comma-separated day numbers to extract, e.g. '1,2,3,4,5'. "
#             "Omit this to auto-extract every day in the month detected from the sheet."
#         )
#     )
# ):
#     # ── Validate file type ───────────────────────────────────────────
#     if not file.filename.endswith((".xlsx", ".xls")):
#         raise HTTPException(
#             status_code=400,
#             detail=f"Invalid file type '{file.filename}'. Only .xlsx or .xls files are accepted."
#         )
 
#     # ── Read file bytes ──────────────────────────────────────────────
#     try:
#         file_bytes = await file.read()
#         if len(file_bytes) == 0:
#             raise HTTPException(status_code=400, detail="Uploaded file is empty.")
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=f"Failed to read uploaded file: {str(e)}")
 
#     # ── Resolve requested days (auto from month, or explicit override) ─
#     try:
#         days_list = resolve_days_list(file_bytes, days)
#     except ValueError as e:
#         raise HTTPException(
#             status_code=400,
#             detail=f"Invalid 'days' parameter: {str(e)}. Expected format: '1,2,3,4,5'"
#         )
 
#     # ── Parse ────────────────────────────────────────────────────────
#     try:
#         result = parse_biometric_excel(file_bytes, days_list)
#         return result
#     except ValueError as e:
#         raise HTTPException(status_code=422, detail=str(e))
#     except Exception as e:
#         raise HTTPException(
#             status_code=500,
#             detail=f"Parsing failed: {str(e)}\n{traceback.format_exc()}"
#         )
 
 
# # ── ENDPOINT 2: Parse from server path (for quick Postman testing) ───
 
# @app.get(
#     "/api/bio/parse-sample",
#     response_model=BiometricParseResponse,
#     tags=["Biometric"],
#     summary="Parse Sample File (Dev/Test Only)",
#     description="""
#     Parses a hardcoded biometric file from the server filesystem.
#     Useful for quick Postman testing without needing to upload a file.
 
#     **Days parameter (auto by default):** same auto-month-detection
#     behavior as `/api/bio/parse` — omit `days` to get every day in the
#     detected month.
 
#     **How to use in Postman:**
#     - Method: GET
#     - URL: http://localhost:8000/api/bio/parse-sample
 
#     **Query params:**
#     - `days` → comma-separated day numbers (optional — omit for full month)
#     - `emp_code` → filter by specific VT ID (optional)
#     """
# )
# async def parse_sample_file(
#     days: Optional[str] = Query(
#         default=None,
#         description="Comma-separated day numbers, e.g. '1,2,3,4,5'. Omit for full auto-detected month."
#     ),
#     emp_code: Optional[str] = Query(
#         default=None,
#         description="Filter by employee VT ID. Example: VT000010349"
#     )
# ):
#     # Hardcoded sample file path (update this to your actual path)
#     SAMPLE_FILE_PATH = "NSE_biometric_data.xlsx"
 
#     # ── Read file ────────────────────────────────────────────────────
#     try:
#         with open(SAMPLE_FILE_PATH, "rb") as f:
#             file_bytes = f.read()
#     except FileNotFoundError:
#         raise HTTPException(
#             status_code=404,
#             detail=f"Sample file not found at '{SAMPLE_FILE_PATH}'. "
#                    f"Use POST /api/bio/parse to upload your file instead."
#         )
 
#     # ── Resolve requested days (auto from month, or explicit override) ─
#     try:
#         days_list = resolve_days_list(file_bytes, days)
#     except ValueError as e:
#         raise HTTPException(
#             status_code=400,
#             detail=f"Invalid 'days' parameter: {str(e)}"
#         )
 
#     # ── Parse ────────────────────────────────────────────────────────
#     try:
#         result = parse_biometric_excel(file_bytes, days_list)
 
#         # Optional: filter by emp_code
#         if emp_code:
#             filtered = [e for e in result.employees if e.emp_code == emp_code.strip()]
#             if not filtered:
#                 raise HTTPException(
#                     status_code=404,
#                     detail=f"Employee '{emp_code}' not found in biometric file."
#                 )
#             result.employees = filtered
#             result.total_employees = len(filtered)
 
#         return result
#     except HTTPException:
#         raise
#     except ValueError as e:
#         raise HTTPException(status_code=422, detail=str(e))
#     except Exception as e:
#         raise HTTPException(
#             status_code=500,
#             detail=f"Parsing failed: {str(e)}\n{traceback.format_exc()}"
#         )
 
 
# # ── ENDPOINT 3: Get single employee attendance ───────────────────────
 
# @app.post(
#     "/api/bio/parse/employee/{emp_code}",
#     response_model=EmployeeRecord,
#     tags=["Biometric"],
#     summary="Get Single Employee Attendance",
#     description="""
#     Upload a biometric file and extract attendance for a specific employee only.
 
#     **Days parameter (auto by default):** omit `days` to auto-extract every
#     day in the month detected from the sheet.
 
#     **How to use in Postman:**
#     - Method: POST
#     - URL: http://localhost:8000/api/bio/parse/employee/VT000010349
#     - Body → form-data:
#         - Key: `file`  Type: File  Value: (select your .xlsx file)
#     """
# )
# async def parse_single_employee(
#     emp_code: str,
#     file: UploadFile = File(...),
#     days: Optional[str] = Query(default=None)
# ):
#     if not file.filename.endswith((".xlsx", ".xls")):
#         raise HTTPException(status_code=400, detail="Only .xlsx or .xls files accepted.")
 
#     file_bytes = await file.read()
 
#     try:
#         days_list = resolve_days_list(file_bytes, days)
#     except ValueError as e:
#         raise HTTPException(status_code=400, detail=f"Invalid days format: {str(e)}")
 
#     try:
#         result = parse_biometric_excel(file_bytes, days_list)
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))
 
#     # Find the specific employee
#     emp_record = next(
#         (e for e in result.employees if e.emp_code == emp_code.strip()),
#         None
#     )
 
#     if not emp_record:
#         raise HTTPException(
#             status_code=404,
#             detail=f"Employee '{emp_code}' not found. "
#                    f"Total employees in file: {result.total_employees}"
#         )
 
#     return emp_record
 
 
# # ─────────────────────────────────────────────────────────────
# # RUN SERVER
# # ─────────────────────────────────────────────────────────────
 
# if __name__ == "__main__":
#     import uvicorn
#     uvicorn.run(
#         "biometric_parser:app",
#         host="0.0.0.0",
#         port=8000,
#         reload=True,         # Auto-reload on file change (dev mode)
#         log_level="info"
#     )



"""
NSE Timesheet Automation System
================================
FastAPI Endpoint: Biometric File Parser — FINAL v3

What this version does (merged from v1 + v2):
  ✅ Auto-detects month & year from "From :01/04/2026 to 30/04/2026" in the sheet
  ✅ Uses calendar.monthrange() to get exact days (28/29/30/31) — no hardcoding
  ✅ NSE Calendar integration — every day enriched with:
       date_label    → "2026-04-03"
       day_name      → "Friday"
       calendar_type → "holiday" / "weekend" / "working_day"
       holiday_name  → "Good Friday" / "Saturday" / None
  ✅ Status WW → shows Saturday / Sunday
  ✅ Status HH → shows exact NSE holiday name
  ✅ All 4 endpoints ready for Postman

Usage:
  uvicorn biometric_parser:app --reload --port 8000
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Query
from pydantic import BaseModel
from typing import Optional
import openpyxl
import pandas as pd
from io import BytesIO
from datetime import date, datetime
import calendar
import re
import traceback
from fastapi.middleware.cors import CORSMiddleware


app = FastAPI(
    title="NSE Timesheet Automation API",
    description="Biometric parser with auto month detection + NSE calendar integration",
    version="3.0.0"
)

origins = [
    "*",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,      # Allowed frontend URLs
    allow_credentials=True,
    allow_methods=["*"],        # GET, POST, PUT, DELETE, etc.
    allow_headers=["*"],        # Allow all headers
)


# ─────────────────────────────────────────────────────────────
# NSE CALENDAR 2026 — Official Trading Holidays
# Source: NSE India website (Trading Holidays 2026)
# ─────────────────────────────────────────────────────────────

NSE_HOLIDAYS_2026 = {
    date(2026, 1, 15): "Municipal Corporation Election - Maharashtra",
    date(2026, 1, 26): "Republic Day",
    date(2026, 3, 3):  "Holi",
    date(2026, 3, 26): "Shri Ram Navami",
    date(2026, 3, 31): "Shri Mahavir Jayanti",
    date(2026, 4, 3):  "Good Friday",
    date(2026, 4, 14): "Dr. Baba Saheb Ambedkar Jayanti",
    date(2026, 5, 1):  "Maharashtra Day",
    date(2026, 5, 28): "Bakri Id",
    date(2026, 6, 26): "Muharram",
    date(2026, 9, 14): "Ganesh Chaturthi",
    date(2026, 10, 2): "Mahatma Gandhi Jayanti",
    date(2026, 10, 20): "Dussehra",
    date(2026, 11, 8):  "Diwali - Laxmi Pujan",
    date(2026, 11, 10): "Diwali - Balipratipada",
    date(2026, 11, 24): "Prakash Gurpurb Sri Guru Nanak Dev",
    date(2026, 12, 25): "Christmas",
}


# ─────────────────────────────────────────────────────────────
# PYDANTIC MODELS
# ─────────────────────────────────────────────────────────────

class DailyAttendance(BaseModel):
    day: int                           # 1, 2, 3 ... 31
    date_label: Optional[str]          # "2026-04-03"
    day_name: Optional[str]            # "Friday"
    calendar_type: Optional[str]       # "working_day" / "weekend" / "holiday"
    holiday_name: Optional[str]        # "Good Friday" / "Saturday" / "Sunday" / None
    shift_code: Optional[str]          # "IO"
    login: Optional[str]               # "09:32" / None
    logout: Optional[str]              # "18:32" / None
    work_hours: Optional[str]          # "9:00" / None
    extra_work_hours: Optional[str]    # "0:30" / None
    status: Optional[str]              # "PP" / "AA" / "WW" / "HH" etc.
    timesheet_value: Optional[float]   # 1.0 / 0.5 / 0.0 / None (blank)


class EmployeeSummary(BaseModel):
    present: int       # P=20
    absent: int        # A=0
    weekend: int       # W=8
    holiday: int       # H=2
    overtime: int      # O=0
    flexi: int         # F=0
    late: int          # L=0
    comp_off: int      # C=0


class EmployeeRecord(BaseModel):
    emp_code: str                       # VT000010349
    emp_name: str                       # Sachin Bijwar
    po_number: str                      # PO-NSEIL OU-26-2027
    summary: EmployeeSummary
    attendance: list[DailyAttendance]


class MonthInfo(BaseModel):
    month: str             # "April"
    month_number: int      # 4
    year: int              # 2026
    days_in_month: int     # 30
    start_date: str        # "01/04/2026"
    end_date: str          # "30/04/2026"


class CalendarSummary(BaseModel):
    working_days: int
    holidays: int
    weekends: int
    holidays_detail: dict  # { "2026-04-03": "Good Friday", ... }


class BiometricParseResponse(BaseModel):
    month_range: str
    month_info: Optional[MonthInfo]
    calendar_summary: CalendarSummary
    total_employees: int
    days_extracted: list[int]
    employees: list[EmployeeRecord]


# ─────────────────────────────────────────────────────────────
# CALENDAR HELPERS
# ─────────────────────────────────────────────────────────────

def get_calendar_info(year: int, month: int, day: int) -> dict:
    """
    Returns NSE calendar metadata for a given date.

    Returns:
        date_label    → "2026-04-03"
        day_name      → "Friday"
        calendar_type → "holiday" / "weekend" / "working_day"
        holiday_name  → "Good Friday" / "Saturday" / "Sunday" / None
    """
    try:
        d = date(year, month, day)
    except ValueError:
        return {
            "date_label": f"{year}-{month:02d}-{day:02d}",
            "day_name": "Unknown",
            "calendar_type": "unknown",
            "holiday_name": None
        }

    day_name = d.strftime("%A")

    # NSE holiday?
    if d in NSE_HOLIDAYS_2026:
        return {
            "date_label": str(d),
            "day_name": day_name,
            "calendar_type": "holiday",
            "holiday_name": NSE_HOLIDAYS_2026[d]
        }

    # Weekend?
    if d.weekday() == 5:
        return {
            "date_label": str(d),
            "day_name": day_name,
            "calendar_type": "weekend",
            "holiday_name": "Saturday"
        }
    if d.weekday() == 6:
        return {
            "date_label": str(d),
            "day_name": day_name,
            "calendar_type": "weekend",
            "holiday_name": "Sunday"
        }

    # Regular working day
    return {
        "date_label": str(d),
        "day_name": day_name,
        "calendar_type": "working_day",
        "holiday_name": None
    }


def build_calendar_summary(year: int, month: int, days: list[int]) -> CalendarSummary:
    """Builds calendar summary counts for the given days."""
    working = holidays = weekends = 0
    holidays_detail = {}

    for day in days:
        info = get_calendar_info(year, month, day)
        if info["calendar_type"] == "working_day":
            working += 1
        elif info["calendar_type"] == "holiday":
            holidays += 1
            holidays_detail[info["date_label"]] = info["holiday_name"]
        elif info["calendar_type"] == "weekend":
            weekends += 1

    return CalendarSummary(
        working_days=working,
        holidays=holidays,
        weekends=weekends,
        holidays_detail=holidays_detail
    )


# ─────────────────────────────────────────────────────────────
# MONTH AUTO-DETECTION
# ─────────────────────────────────────────────────────────────

def get_month_details(file_bytes: bytes) -> dict:
    """
    Reads "From :01/04/2026 to 30/04/2026" from the Excel sheet.
    Returns structured month info dict.
    """
    df = pd.read_excel(BytesIO(file_bytes), header=None, engine="openpyxl")

    date_text = None
    for row in df.itertuples(index=False):
        for cell in row:
            if pd.notna(cell) and "From" in str(cell) and "to" in str(cell):
                date_text = str(cell)
                break
        if date_text:
            break

    if not date_text:
        raise ValueError("Could not find attendance period in Excel. Expected 'From :DD/MM/YYYY to DD/MM/YYYY'.")

    match = re.search(r'(\d{2}/\d{2}/\d{4})\s*to\s*(\d{2}/\d{2}/\d{4})', date_text)
    if not match:
        raise ValueError(f"Date format not recognized in: '{date_text}'")

    start_dt = datetime.strptime(match.group(1), "%d/%m/%Y")
    end_dt   = datetime.strptime(match.group(2), "%d/%m/%Y")

    return {
        "month":         start_dt.strftime("%B"),
        "month_number":  start_dt.month,
        "year":          start_dt.year,
        "days_in_month": calendar.monthrange(start_dt.year, start_dt.month)[1],
        "start_date":    start_dt.strftime("%d/%m/%Y"),
        "end_date":      end_dt.strftime("%d/%m/%Y")
    }


def resolve_days_list(file_bytes: bytes, days: Optional[str]) -> tuple[list[int], dict]:
    """
    Determines which days to extract.

    - days=None  → auto detect month from sheet → return all days (1..30/31)
    - days='1,2,3' → parse and validate explicit override

    Returns:
        (days_list, month_details_dict)
    """
    month_details = get_month_details(file_bytes)  # always detect month

    if not days or not days.strip():
        # Auto: use all days in month
        days_list = list(range(1, month_details["days_in_month"] + 1))
        return days_list, month_details

    # Explicit override
    try:
        days_list = list(set([int(d.strip()) for d in days.split(",") if d.strip()]))
    except ValueError:
        raise ValueError("Invalid days format. Use comma-separated numbers: '1,2,3,4,5'")

    if not days_list:
        raise ValueError("Empty days list provided.")
    if any(d < 1 or d > 31 for d in days_list):
        raise ValueError("Day numbers must be between 1 and 31.")

    return sorted(days_list), month_details


# ─────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────

def clean_val(val) -> Optional[str]:
    """Returns None for empty/dash values, clean string otherwise."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("-", "", "None") else s


def parse_summary(summary_values: list[str]) -> EmployeeSummary:
    """Parses ['P=20', 'A=0', 'W=8', ...] into EmployeeSummary."""
    m = {}
    for token in summary_values:
        if "=" in token:
            key, _, val = token.partition("=")
            try:
                m[key.strip().upper()] = int(val.strip())
            except ValueError:
                m[key.strip().upper()] = 0
    return EmployeeSummary(
        present  = m.get("P", 0),
        absent   = m.get("A", 0),
        weekend  = m.get("W", 0),
        holiday  = m.get("H", 0),
        overtime = m.get("O", 0),
        flexi    = m.get("F", 0),
        late     = m.get("L", 0),
        comp_off = m.get("C", 0),
    )


def parse_work_hours(wh_str: str) -> Optional[float]:
    """Converts '9:30' → 9.5 decimal hours. Returns None if unparseable."""
    if not wh_str:
        return None
    try:
        parts = wh_str.split(":")
        return round(int(parts[0]) + int(parts[1]) / 60, 2)
    except Exception:
        return None


def map_status_to_timesheet_value(
    status: Optional[str],
    work_hours: Optional[str]
) -> Optional[float]:
    """
    Maps biometric status → timesheet cell value.
      1.0  → Full day present  (PP / PP* with >= 9 hrs)
      0.5  → Half day          (PA / PA*)
      0.0  → Absent            (AA)
      None → Weekend / Holiday (WW / HH — blank in timesheet)
    """
    if not status:
        return None

    s = status.upper().strip()

    if s in ("PP", "PP*"):
        hours = parse_work_hours(work_hours)
        if hours is not None:
            if hours >= 9.0:  return 1.0
            elif hours >= 4.5: return 0.5
            else:              return 0.0
        return 1.0  # trust status code if hours can't be parsed

    if s in ("PA", "PA*"): return 0.5
    if s == "AA":          return 0.0
    if s in ("WW", "WW*", "WW#"):  return None   # weekend → blank
    if s in ("HH", "HH#"):         return None   # holiday → blank
    if s == "HH*":                 return 1.0    # worked on holiday → 1

    return None  # unknown status


# ─────────────────────────────────────────────────────────────
# CORE PARSER
# ─────────────────────────────────────────────────────────────

def parse_biometric_excel(
    file_bytes: bytes,
    days_filter: list[int],
    month_details: dict
) -> BiometricParseResponse:
    """
    Main parser. Reads NSE biometric Excel and returns structured
    attendance enriched with NSE calendar info.

    Args:
        file_bytes    : Raw .xlsx bytes
        days_filter   : List of day numbers to extract
        month_details : Dict from get_month_details() with year/month info
    """
    year  = month_details["year"]
    month = month_details["month_number"]

    # ── Build calendar lookup for all requested days ─────────────────
    # { day_int → calendar_info_dict }
    cal_map = {day: get_calendar_info(year, month, day) for day in days_filter}

    # ── Build calendar summary ───────────────────────────────────────
    cal_summary = build_calendar_summary(year, month, days_filter)

    # ── Open Excel ───────────────────────────────────────────────────
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active   # Sheet: MusterInOutReport

    # ── Row 5: month range string ────────────────────────────────────
    row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    month_range = str(row5[0]) if row5[0] else "Unknown"

    # ── Row 6: date header — map day number → column index (0-based) ─
    header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    date_col_map = {}
    for col_idx, cell_val in enumerate(header_row):
        if isinstance(cell_val, int) and 1 <= cell_val <= 31:
            if cell_val not in date_col_map:
                date_col_map[cell_val] = col_idx

    # Validate all requested days exist in the file
    invalid_days = [d for d in days_filter if d not in date_col_map]
    if invalid_days:
        raise ValueError(
            f"Days {invalid_days} not found in file header. "
            f"Available: {sorted(date_col_map.keys())}"
        )

    # ── Bulk read all data rows at once (no N+1) ─────────────────────
    ROW_TYPES = ["ShiftCode", "Login", "Logout", "WorkHours", "ExtraWorkHours", "Status", "Summary"]
    all_rows = list(ws.iter_rows(min_row=7, values_only=True))
    wb.close()

    # ── Parse employee blocks (7 rows each) ─────────────────────────
    employees = []
    i = 0

    while i < len(all_rows):
        shift_row = all_rows[i]
        emp_code  = shift_row[0]

        # Each employee block starts with VT code in column 0
        if not emp_code or not str(emp_code).startswith("VT"):
            i += 1
            continue

        if i + 7 > len(all_rows):
            break

        block   = all_rows[i: i + 7]
        row_map = {rt: block[j] for j, rt in enumerate(ROW_TYPES)}

        emp_name  = str(shift_row[1]).strip() if shift_row[1] else ""
        po_number = str(shift_row[2]).strip() if shift_row[2] else ""

        # Parse summary (P=20, A=0, W=8, H=2...)
        summary_tokens = [
            str(v).strip() for v in row_map["Summary"]
            if v is not None and str(v).strip() not in ("", "None")
        ]
        summary = parse_summary(summary_tokens)

        # Build daily attendance — one entry per day, enriched with calendar
        attendance = []
        for day in sorted(days_filter):
            col = date_col_map[day]
            cal = cal_map[day]   # calendar info for this date

            raw_shift  = clean_val(row_map["ShiftCode"][col])
            raw_login  = clean_val(row_map["Login"][col])
            raw_logout = clean_val(row_map["Logout"][col])
            raw_wh     = clean_val(row_map["WorkHours"][col])
            raw_ewh    = clean_val(row_map["ExtraWorkHours"][col])
            raw_status = clean_val(row_map["Status"][col])

            ts_value = map_status_to_timesheet_value(raw_status, raw_wh)

            attendance.append(DailyAttendance(
                day=day,
                # ── Calendar fields (NEW) ──────────────────────────
                date_label=cal["date_label"],
                day_name=cal["day_name"],
                calendar_type=cal["calendar_type"],
                holiday_name=cal["holiday_name"],
                # ── Biometric fields ──────────────────────────────
                shift_code=raw_shift,
                login=raw_login,
                logout=raw_logout,
                work_hours=raw_wh,
                extra_work_hours=raw_ewh,
                status=raw_status,
                timesheet_value=ts_value
            ))

        employees.append(EmployeeRecord(
            emp_code=str(emp_code).strip(),
            emp_name=emp_name,
            po_number=po_number,
            summary=summary,
            attendance=attendance
        ))

        i += 7

    return BiometricParseResponse(
        month_range=month_range,
        month_info=MonthInfo(**month_details),
        calendar_summary=cal_summary,
        total_employees=len(employees),
        days_extracted=sorted(days_filter),
        employees=employees
    )


# ─────────────────────────────────────────────────────────────
# API ENDPOINTS
# ─────────────────────────────────────────────────────────────

@app.get("/", tags=["Health"])
async def root():
    return {
        "service": "NSE Timesheet Automation API",
        "version": "3.0.0",
        "endpoints": {
            "upload_and_parse":  "POST /api/bio/parse",
            "parse_sample":      "GET  /api/bio/parse-sample",
            "single_employee":   "POST /api/bio/parse/employee/{emp_code}",
            "nse_calendar":      "GET  /api/calendar/{year}/{month}",
            "swagger_docs":      "GET  /docs"
        }
    }


@app.get("/health", tags=["Health"])
async def health_check():
    return {"status": "ok"}


# ── ENDPOINT 1: Upload & Parse ───────────────────────────────────────

@app.post(
    "/api/bio/parse",
    response_model=BiometricParseResponse,
    tags=["Biometric"],
    summary="Upload & Parse Biometric File (auto month detection)",
    description="""
Upload an NSE biometric .xlsx file. The API will:
1. Auto-detect the month & year from the sheet ("From :01/04/2026 to 30/04/2026")
2. Use `calendar.monthrange()` to determine the correct number of days (28/29/30/31)
3. Enrich every day with NSE calendar info (holiday name, day name, calendar type)

**Postman Setup:**
- Method: `POST`
- URL: `http://localhost:8000/api/bio/parse`
- Body → `form-data`:
  - Key: `file` | Type: `File` | Value: *(select .xlsx)*
  - Key: `days` | Type: `Text` | Value: `1,2,3,4,5` *(optional — omit for full month)*

**days param:**
- Omit → auto extracts all days of the detected month
- `1,2,3,4,5` → extracts only those specific days
    """
)
async def parse_biometric_file(
    file: UploadFile = File(..., description="NSE biometric .xlsx file"),
    days: Optional[str] = Query(
        default=None,
        description="Comma-separated day numbers e.g. '1,2,3,4,5'. Omit for full auto-detected month."
    )
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file '{file.filename}'. Only .xlsx or .xls accepted."
        )

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        days_list, month_details = resolve_days_list(file_bytes, days)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    try:
        return parse_biometric_excel(file_bytes, days_list, month_details)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parsing error: {str(e)}\n{traceback.format_exc()}")


# ── ENDPOINT 2: Quick test with server-side sample file ──────────────

@app.get(
    "/api/bio/parse-sample",
    response_model=BiometricParseResponse,
    tags=["Biometric"],
    summary="Parse Sample File (Dev/Test — no upload needed)",
    description="""
Parses the biometric file from the server path — no upload needed.
Same auto month detection + NSE calendar enrichment as the POST endpoint.

**Postman Setup:**
- Method: `GET`
- URL: `http://localhost:8000/api/bio/parse-sample`
- Optional params:
  - `?days=1,2,3,4,5` → specific days only
  - `?emp_code=VT000010349` → filter single employee
    """
)
async def parse_sample_file(
    days: Optional[str] = Query(default=None),
    emp_code: Optional[str] = Query(default=None, description="Filter by VT ID e.g. VT000010349")
):
    SAMPLE_FILE = "NSE_biometric_data.xlsx"

    try:
        with open(SAMPLE_FILE, "rb") as f:
            file_bytes = f.read()
    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail=f"Sample file not found at '{SAMPLE_FILE}'. Use POST /api/bio/parse to upload."
        )

    try:
        days_list, month_details = resolve_days_list(file_bytes, days)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    try:
        result = parse_biometric_excel(file_bytes, days_list, month_details)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if emp_code:
        filtered = [e for e in result.employees if e.emp_code == emp_code.strip()]
        if not filtered:
            raise HTTPException(
                status_code=404,
                detail=f"Employee '{emp_code}' not found. Total in file: {result.total_employees}"
            )
        result.employees = filtered
        result.total_employees = len(filtered)

    return result


# ── ENDPOINT 3: Single employee ──────────────────────────────────────

@app.post(
    "/api/bio/parse/employee/{emp_code}",
    response_model=EmployeeRecord,
    tags=["Biometric"],
    summary="Get Single Employee Attendance",
    description="""
Upload biometric file and extract attendance for one specific employee.

**Postman Setup:**
- Method: `POST`
- URL: `http://localhost:8000/api/bio/parse/employee/VT000010349`
- Body → `form-data`:
  - Key: `file` | Type: `File` | Value: *(select .xlsx)*
  - Key: `days` | Type: `Text` | Value: `1,2,3,4,5` *(optional)*
    """
)
async def parse_single_employee(
    emp_code: str,
    file: UploadFile = File(...),
    days: Optional[str] = Query(default=None)
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls accepted.")

    file_bytes = await file.read()

    try:
        days_list, month_details = resolve_days_list(file_bytes, days)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    try:
        result = parse_biometric_excel(file_bytes, days_list, month_details)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    emp = next((e for e in result.employees if e.emp_code == emp_code.strip()), None)
    if not emp:
        raise HTTPException(
            status_code=404,
            detail=f"Employee '{emp_code}' not found. Total in file: {result.total_employees}"
        )
    return emp


# ── ENDPOINT 4: View NSE Calendar for any month ──────────────────────

@app.get(
    "/api/calendar/{year}/{month}",
    tags=["Calendar"],
    summary="Get NSE Calendar for a Month",
    description="""
Returns full NSE calendar for any month with working days, holidays (with names), and weekends.

**Postman Setup:**
- Method: `GET`
- URL: `http://localhost:8000/api/calendar/2026/4`
    """
)
async def get_nse_calendar(year: int, month: int):
    if not (2020 <= year <= 2030):
        raise HTTPException(status_code=400, detail="Year must be between 2020 and 2030")
    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Month must be between 1 and 12")

    _, days_in_month = calendar.monthrange(year, month)

    cal_data = []
    working = holidays = weekends = 0

    for day in range(1, days_in_month + 1):
        info = get_calendar_info(year, month, day)
        cal_data.append({"day": day, **info})
        if info["calendar_type"] == "working_day": working += 1
        elif info["calendar_type"] == "holiday":   holidays += 1
        elif info["calendar_type"] == "weekend":   weekends += 1

    return {
        "year": year,
        "month": month,
        "month_name": date(year, month, 1).strftime("%B"),
        "total_days": days_in_month,
        "summary": {
            "working_days": working,
            "holidays": holidays,
            "weekends": weekends,
        },
        "calendar": cal_data
    }


# ─────────────────────────────────────────────────────────────
# RUN
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("biometric_parser:app", host="0.0.0.0", port=8000, reload=True)